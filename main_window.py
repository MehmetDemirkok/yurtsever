from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QLineEdit, QDateEdit, 
    QTableWidget, QTableWidgetItem, QComboBox,
    QMessageBox, QHeaderView, QFileDialog, QDialog, QTextBrowser
)
from PyQt5.QtCore import Qt, QDate, pyqtSignal
from PyQt5.QtGui import QFont, QDoubleValidator
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from database import Database
from stay_model import StayModel
from helpers import validate_dates, format_currency
import pandas as pd
from datetime import datetime
from unidecode import unidecode
from typing import Any, Optional
import re

class MainWindow(QMainWindow):
    # Define a signal for report generation completion
    report_generated = pyqtSignal() # This signal is used to show a message after report generation

    # Define required Excel column names as a class attribute for consistent use
    _REQUIRED_EXCEL_COLUMNS_MAP = {
        'adi_soyadi': 'Adı Soyadı',
        'unvan': 'Unvan',
        'ulke': 'Ülke',
        'sehir': 'Şehir',
        'giris_tarihi': 'Giriş Tarihi',
        'cikis_tarihi': 'Çıkış Tarihi',
        'oda_tipi': 'Oda Tipi',
        'otel_alis_fiyati': 'Otel Alış Fiyatı',
        'kurum_cari': 'Kurum / Cari',
        'otel_adi': 'Otel Adı',
        'otel_satis_fiyati': 'Otel Satış Fiyatı'
    }

    def __init__(self):
        super().__init__()
        self.db = Database()
        self.stay_model = StayModel(self.db)
        self.current_sort_column = -1
        self.current_sort_order = Qt.AscendingOrder
        self.init_ui()
        
    @staticmethod
    def normalize_column_name(col_name: Any) -> Optional[str]:
        """Normalizes column names for consistent matching."""
        if not isinstance(col_name, str):
            print(f"Normalize column name: Input is not a string: {col_name}") # Debugging
            return None
        
        original_col_name = col_name # Debugging
        col_str = col_name.strip()
        
        # Replace various whitespace characters (including non-breaking spaces) with a single space
        col_str = re.sub(r'\s+', ' ', col_str) # Replace all whitespace sequences with a single space
        col_str = col_str.replace('\xa0', ' ') # Replace non-breaking space explicitly
        
        # Define a mapping for Turkish characters to ASCII equivalents
        turkish_map = {
            'ç': 'c', 'Ç': 'C',
            'ğ': 'g', 'Ğ': 'G',
            'ı': 'i', 'İ': 'I',
            'ö': 'o', 'Ö': 'O',
            'ş': 's', 'Ş': 'S',
            'ü': 'u', 'Ü': 'U',
        }
        
        # Apply manual mapping first
        for turkish_char, ascii_char in turkish_map.items():
            col_str = col_str.replace(turkish_char, ascii_char)
        
        # Then use unidecode for any other non-ASCII characters
        # This might be redundant if turkish_map covers all, but safe to keep.
        try:
            from unidecode import unidecode
            col_str = unidecode(col_str)
        except ImportError:
            print("Warning: unidecode not found. Ensure it's installed for full normalization.")

        
        # Convert to lowercase
        col_str = col_str.lower()
        
        # Replace non-alphanumeric characters (except underscore) with a single underscore
        col_str = re.sub(r'[^a-z0-9_]+', '_', col_str) # Allow underscore, replace others
        
        # Replace multiple underscores with a single underscore
        col_str = re.sub(r'_+', '_', col_str)
        
        # Remove leading/trailing underscores if any
        col_str = col_str.strip('_')

        print(f"Normalize column name: Original: '{original_col_name}' -> Normalized: '{col_str}'") # Debugging
        return col_str

    def init_ui(self):
        """Initialize the user interface."""
        self.setWindowTitle('Otel Konaklama Puantaj Sistemi')
        self.setMinimumSize(1000, 600)
        
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # Create input form
        form_layout = QVBoxLayout()  # Changed to QVBoxLayout for better organization
        
        # First row - Name and Title
        name_title_layout = QHBoxLayout()
        
        # Guest name input
        self.guest_name_input = QLineEdit()
        self.guest_name_input.setPlaceholderText('Adı Soyadı')
        name_title_layout.addWidget(self.guest_name_input)
        
        # Guest title input
        self.guest_title_input = QLineEdit()
        self.guest_title_input.setPlaceholderText('Unvanı')
        name_title_layout.addWidget(self.guest_title_input)

        # Company Name input
        self.company_name_input = QLineEdit()
        self.company_name_input.setPlaceholderText('Kurum / Cari')
        name_title_layout.addWidget(self.company_name_input)
        
        form_layout.addLayout(name_title_layout)
        
        # Second row - Country and City
        location_layout = QHBoxLayout()
        
        # Country input
        self.country_input = QLineEdit()
        self.country_input.setPlaceholderText('Ülke')
        location_layout.addWidget(self.country_input)
        
        # City input
        self.city_input = QLineEdit()
        self.city_input.setPlaceholderText('Şehir')
        location_layout.addWidget(self.city_input)
        
        form_layout.addLayout(location_layout)
        
        # Third row - Dates and Room Type
        dates_room_layout = QHBoxLayout()
        
        # Check-in date input
        self.check_in_date = QDateEdit()
        self.check_in_date.setDate(QDate.currentDate())
        self.check_in_date.setCalendarPopup(True)
        self.check_in_date.setDisplayFormat("dd.MM.yyyy")
        dates_room_layout.addWidget(QLabel("Giriş Tarihi:"))
        dates_room_layout.addWidget(self.check_in_date)
        
        # Check-out date input
        self.check_out_date = QDateEdit()
        self.check_out_date.setDate(QDate.currentDate().addDays(1))
        self.check_out_date.setCalendarPopup(True)
        self.check_out_date.setDisplayFormat("dd.MM.yyyy")
        dates_room_layout.addWidget(QLabel("Çıkış Tarihi:"))
        dates_room_layout.addWidget(self.check_out_date)
        
        # Room type input
        self.room_type_combo = QComboBox()
        self.room_type_combo.addItems(['Single Oda', 'Double Oda', 'Triple Oda', 'Suit Oda', 'Aile Odası'])
        dates_room_layout.addWidget(QLabel("Oda Tipi:"))
        dates_room_layout.addWidget(self.room_type_combo)
        
        form_layout.addLayout(dates_room_layout)
        
        # All action buttons and nightly rate input in one horizontal layout
        action_buttons_and_rate_layout = QHBoxLayout()
        
        # Nightly rate input
        self.nightly_rate_input = QLineEdit()
        self.nightly_rate_input.setPlaceholderText('Otel Alış Fiyatı')
        self.nightly_rate_input.setValidator(QDoubleValidator(0.00, 999999.99, 2))
        action_buttons_and_rate_layout.addWidget(self.nightly_rate_input)

        # Hotel Sale Price input
        self.hotel_sale_price_input = QLineEdit()
        self.hotel_sale_price_input.setPlaceholderText('Otel Satış Fiyatı')
        self.hotel_sale_price_input.setValidator(QDoubleValidator(0.00, 999999.99, 2))
        action_buttons_and_rate_layout.addWidget(self.hotel_sale_price_input)

        # Hotel Name input
        self.hotel_name_input = QLineEdit()
        self.hotel_name_input.setPlaceholderText('Otel Adı')
        action_buttons_and_rate_layout.addWidget(self.hotel_name_input)
        
        # Add button
        self.add_button = QPushButton('Ekle')
        self.add_button.setObjectName('add_button')
        self.add_button.clicked.connect(self.add_stay)
        action_buttons_and_rate_layout.addWidget(self.add_button)
        
        # Edit selected button
        self.edit_selected_button = QPushButton('Seçili Kaydı Düzenle')
        self.edit_selected_button.clicked.connect(self.edit_selected_stay)
        action_buttons_and_rate_layout.addWidget(self.edit_selected_button)

        # Delete selected button
        self.delete_selected_button = QPushButton('Seçili Kaydı Sil')
        self.delete_selected_button.clicked.connect(self.delete_selected_stay)
        action_buttons_and_rate_layout.addWidget(self.delete_selected_button)

        form_layout.addLayout(action_buttons_and_rate_layout)
        
        # Connect returnPressed signal to add_button's clicked signal
        self.guest_name_input.returnPressed.connect(self.add_button.click)
        self.guest_title_input.returnPressed.connect(self.add_button.click)
        self.country_input.returnPressed.connect(self.add_button.click)
        self.city_input.returnPressed.connect(self.add_button.click)
        self.nightly_rate_input.returnPressed.connect(self.add_button.click)
        self.company_name_input.returnPressed.connect(self.add_button.click)
        self.hotel_name_input.returnPressed.connect(self.add_button.click)
        self.hotel_sale_price_input.returnPressed.connect(self.add_button.click)
        
        main_layout.addLayout(form_layout)
        
        # Create table
        self.table = QTableWidget()
        self.table.setColumnCount(14)  # Changed column count from 12 to 14
        self.table.setHorizontalHeaderLabels([
            'ID', 'Adı Soyadı', 'Unvan', 'Kurum / Cari', 'Ülke', 'Şehir',
            'Giriş Tarihi', 'Çıkış Tarihi', 'Oda Tipi', 'Otel Alış Fiyatı', 'Otel Alış Toplam Ücreti', 'Otel Adı', 'Otel Satış Fiyatı', 'Toplam Satış Tutarı'
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.horizontalHeader().sectionClicked.connect(self.sort_data)
        main_layout.addWidget(self.table)
        
        # Create filter controls
        filter_layout = QHBoxLayout()
        
        self.filter_guest_name_input = QLineEdit()
        self.filter_guest_name_input.setPlaceholderText('Misafir Adı Soyadı Filtrele')
        self.filter_guest_name_input.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_guest_name_input)
        
        self.filter_country_input = QLineEdit()
        self.filter_country_input.setPlaceholderText('Ülke Filtrele')
        self.filter_country_input.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_country_input)
        
        self.filter_city_input = QLineEdit()
        self.filter_city_input.setPlaceholderText('Şehir Filtrele')
        self.filter_city_input.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_city_input)

        self.filter_room_type_combo = QComboBox()
        self.filter_room_type_combo.addItems(['Tümü', 'Single Oda', 'Double Oda', 'Triple Oda', 'Suit Oda', 'Aile Odası'])
        self.filter_room_type_combo.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_room_type_combo)

        main_layout.addLayout(filter_layout)
        
        # Create report and import controls
        control_layout = QHBoxLayout()
        
        # Statistics Report button
        stats_button = QPushButton('Puantaj Raporu')
        stats_button.clicked.connect(self.generate_puantaj_report)
        control_layout.addWidget(stats_button)

        # Import Excel button
        import_excel_button = QPushButton('Excel\'den İçe Aktar')
        import_excel_button.clicked.connect(self.import_excel)
        control_layout.addWidget(import_excel_button)

        # Export Excel button
        export_excel_button = QPushButton('Excel\'e Aktar')
        export_excel_button.clicked.connect(self.export_excel)
        control_layout.addWidget(export_excel_button)
        
        # Download Excel Template button
        download_template_button = QPushButton('Excel Şablonu İndir')
        download_template_button.clicked.connect(self.download_excel_template)
        control_layout.addWidget(download_template_button)

        main_layout.addLayout(control_layout)

        # Load initial data
        self.load_stays()

        # Connect selection change signal to enable/disable edit/delete buttons
        self.table.selectionModel().selectionChanged.connect(self.update_action_buttons_state)
        self.update_action_buttons_state() # Initial state

    def add_stay(self):
        """Add a new stay record to the database from input fields."""
        guest_name = self.guest_name_input.text().strip()
        guest_title = self.guest_title_input.text().strip()
        company_name = self.company_name_input.text().strip()
        country = self.country_input.text().strip()
        city = self.city_input.text().strip()
        check_in_date = self.check_in_date.date().toString("yyyy-MM-dd")
        check_out_date = self.check_out_date.date().toString("yyyy-MM-dd")
        room_type = self.room_type_combo.currentText()
        hotel_purchase_price_text = self.nightly_rate_input.text().strip()
        hotel_name = self.hotel_name_input.text().strip()
        hotel_sale_price_text = self.hotel_sale_price_input.text().strip()

        if not all([guest_name, guest_title, country, city, check_in_date, check_out_date, hotel_purchase_price_text, hotel_name, hotel_sale_price_text]):
            QMessageBox.warning(self, "Eksik Bilgi", "Lütfen tüm alanları doldurun.")
            return
        
        try:
            hotel_purchase_price = float(hotel_purchase_price_text.replace("₺", "").replace(".", "").replace(",", "."))
            hotel_sale_price = float(hotel_sale_price_text.replace("₺", "").replace(".", "").replace(",", "."))
        except ValueError:
            QMessageBox.warning(self, "Geçersiz Giriş", "Otel Alış Fiyatı ve Otel Satış Fiyatı geçerli bir sayı olmalıdır.")
            return

        if not validate_dates(check_in_date, check_out_date):
            QMessageBox.warning(self, "Geçersiz Tarih", "Çıkış tarihi, giriş tarihinden sonra olmalıdır.")
            return

        try:
            self.stay_model.create_stay(
                guest_name, guest_title, company_name, country, city,
                check_in_date, check_out_date, room_type,
                hotel_purchase_price, hotel_name, hotel_sale_price
            )
            self.load_stays() # Refresh table
            self.clear_inputs()
            QMessageBox.information(self, "Başarılı", "Konaklama kaydı başarıyla eklendi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Konaklama kaydı eklenirken bir hata oluştu: {e}")

    def load_stays(self, guest_name: str = None, country: str = None, city: str = None, room_type: str = None, sort_column_index: int = None, sort_order: Qt.SortOrder = Qt.AscendingOrder):
        """Load stay records into the table, with optional filtering and sorting."""
        try:
            # Get stays from model
            sort_column_text = None
            if sort_column_index is not None and sort_column_index != -1 and self.table.horizontalHeaderItem(sort_column_index):
                sort_column_text = self.table.horizontalHeaderItem(sort_column_index).text()
            stays = self.stay_model.get_all_stays(guest_name, country, city, room_type, sort_column_text, "ASC" if sort_order == Qt.AscendingOrder else "DESC")
            
            self.table.setRowCount(len(stays))
            for row, stay in enumerate(stays):
                 self.table.setItem(row, 0, QTableWidgetItem(str(stay['id'])))
                 self.table.setItem(row, 1, QTableWidgetItem(stay['guest_name']))
                 self.table.setItem(row, 2, QTableWidgetItem(stay['guest_title']))
                 self.table.setItem(row, 3, QTableWidgetItem(stay['company_name']))
                 self.table.setItem(row, 4, QTableWidgetItem(stay['country']))
                 self.table.setItem(row, 5, QTableWidgetItem(stay['city']))
                 self.table.setItem(row, 6, QTableWidgetItem(stay['check_in_date']))
                 self.table.setItem(row, 7, QTableWidgetItem(stay['check_out_date']))
                 self.table.setItem(row, 8, QTableWidgetItem(stay['room_type']))
                 self.table.setItem(row, 9, QTableWidgetItem(format_currency(stay['hotel_purchase_price'])))
                 self.table.setItem(row, 10, QTableWidgetItem(format_currency(stay['hotel_purchase_total_amount'])))
                 self.table.setItem(row, 11, QTableWidgetItem(stay['hotel_name']))
                 self.table.setItem(row, 12, QTableWidgetItem(format_currency(stay['hotel_sale_price'])))
                 self.table.setItem(row, 13, QTableWidgetItem(format_currency(stay['total_sale_amount'])))
                  
            # Ensure sort_column_index is an integer before passing to sortItems
            display_sort_column = sort_column_index if sort_column_index is not None and sort_column_index != -1 else 0
            self.table.sortItems(display_sort_column, sort_order) # Apply visual sort

            self.update_action_buttons_state() # Update button state after loading

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kayıtlar yüklenirken bir hata oluştu: {e}")

    def clear_inputs(self):
        """Clear all input fields."""
        self.guest_name_input.clear()
        self.guest_title_input.clear()
        self.company_name_input.clear()
        self.country_input.clear()
        self.city_input.clear()
        self.nightly_rate_input.clear()
        self.hotel_name_input.clear()
        self.hotel_sale_price_input.clear()
        self.check_in_date.setDate(QDate.currentDate())
        self.check_out_date.setDate(QDate.currentDate().addDays(1))
        self.room_type_combo.setCurrentIndex(0)
        
        # Reset add button to its original state
        self.add_button.setText('Ekle')
        self.add_button.setObjectName('add_button')
        try:
            self.add_button.clicked.disconnect()
        except TypeError: # Signal not connected
            pass
        self.add_button.clicked.connect(self.add_stay)

        self.update_action_buttons_state()

    def update_action_buttons_state(self):
        """Enable or disable edit/delete buttons based on table selection."""
        is_selected = len(self.table.selectionModel().selectedRows()) > 0
        self.edit_selected_button.setEnabled(is_selected)
        self.delete_selected_button.setEnabled(is_selected)

    def edit_stay(self, row):
        """Populate input fields with data from the selected row for editing."""
        try:
            stay_id = int(self.table.item(row, 0).text())
            stay = self.stay_model.get_stay_by_id(stay_id)

            if stay:
                self.guest_name_input.setText(stay['guest_name'])
                self.guest_title_input.setText(stay['guest_title'])
                self.company_name_input.setText(stay['company_name'])
                self.country_input.setText(stay['country'])
                self.city_input.setText(stay['city'])
                self.check_in_date.setDate(QDate.fromString(stay['check_in_date'], 'yyyy-MM-dd'))
                self.check_out_date.setDate(QDate.fromString(stay['check_out_date'], 'yyyy-MM-dd'))
                self.room_type_combo.setCurrentText(stay['room_type'])
                self.nightly_rate_input.setText(str(stay['hotel_purchase_price']))
                self.hotel_sale_price_input.setText(str(stay['hotel_sale_price']))
                self.hotel_name_input.setText(stay['hotel_name'])

                # Change add button to save button
                self.add_button.setText('Kaydet')
                self.add_button.setObjectName('save_button') # Change object name to distinguish
                # Disconnect old slot if connected and connect new one
                try:
                    self.add_button.clicked.disconnect()
                except TypeError: # Signal not connected
                    pass
                self.add_button.clicked.connect(lambda: self.save_edit(stay_id))

                # Enable edit/delete buttons
                self.edit_selected_button.setEnabled(True)
                self.delete_selected_button.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kaydı düzenlerken bir hata oluştu: {e}")

    def edit_selected_stay(self):
        """Edit the currently selected stay record."""
        selected_rows = self.table.selectionModel().selectedRows()
        if selected_rows:
            row = selected_rows[0].row()
            self.edit_stay(row)
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen düzenlemek için bir kayıt seçin.")

    def delete_stay(self, row):
        """Delete a stay record from the database."""
        stay_id = int(self.table.item(row, 0).text())
        reply = QMessageBox.question(
            self, 
            'Kaydı Sil',
            f'ID: {stay_id} olan kaydı silmek istediğinizden emin misiniz?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                self.stay_model.delete_stay(stay_id)
                self.load_stays()
                self.clear_inputs()
                QMessageBox.information(self, "Başarılı", "Kayıt başarıyla silindi.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kayıt silinirken bir hata oluştu: {e}")

    def delete_selected_stay(self):
        """Delete the currently selected stay record."""
        selected_rows = self.table.selectionModel().selectedRows()
        if selected_rows:
            row = selected_rows[0].row()
            self.delete_stay(row)
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir kayıt seçin.")

    def save_edit(self, stay_id):
        """Save changes to an existing stay record."""
        guest_name = self.guest_name_input.text().strip()
        guest_title = self.guest_title_input.text().strip()
        company_name = self.company_name_input.text().strip()
        country = self.country_input.text().strip()
        city = self.city_input.text().strip()
        check_in_date = self.check_in_date.date().toString("yyyy-MM-dd")
        check_out_date = self.check_out_date.date().toString("yyyy-MM-dd")
        room_type = self.room_type_combo.currentText()
        hotel_purchase_price_text = self.nightly_rate_input.text().strip()
        hotel_sale_price_text = self.hotel_sale_price_input.text().strip()
        hotel_name = self.hotel_name_input.text().strip()

        if not all([guest_name, guest_title, country, city, check_in_date, check_out_date, hotel_purchase_price_text, hotel_name, hotel_sale_price_text]):
            QMessageBox.warning(self, "Eksik Bilgi", "Lütfen tüm alanları doldurun.")
            return
        
        try:
            hotel_purchase_price = float(hotel_purchase_price_text.replace("₺", "").replace(".", "").replace(",", "."))
            hotel_sale_price = float(hotel_sale_price_text.replace("₺", "").replace(".", "").replace(",", "."))
        except ValueError:
            QMessageBox.warning(self, "Geçersiz Giriş", "Otel Alış Fiyatı ve Otel Satış Fiyatı geçerli bir sayı olmalıdır.")
            return

        if not validate_dates(check_in_date, check_out_date):
            QMessageBox.warning(self, "Geçersiz Tarih", "Çıkış tarihi, giriş tarihinden sonra olmalıdır.")
            return

        try:
            success = self.stay_model.update_stay(
                stay_id,
                guest_name=guest_name,
                guest_title=guest_title,
                company_name=company_name,
                country=country,
                city=city,
                check_in_date=check_in_date,
                check_out_date=check_out_date,
                room_type=room_type,
                hotel_purchase_price=hotel_purchase_price,
                hotel_sale_price=hotel_sale_price,
                hotel_name=hotel_name
            )
            if success:
                self.load_stays()
                self.clear_inputs()
                # Revert save button to add button
                self.add_button.setText('Ekle')
                self.add_button.setObjectName('add_button')
                try:
                    self.add_button.clicked.disconnect()
                except TypeError: # Signal not connected
                    pass
                self.add_button.clicked.connect(self.add_stay)
                QMessageBox.information(self, "Başarılı", "Kayıt başarıyla güncellendi.")
            else:
                QMessageBox.warning(self, "Uyarı", "Kayıt güncellenemedi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kayıt güncellenirken bir hata oluştu: {e}")
    
    def import_excel(self):
        """Import data from Excel file."""
        try:
            # Open file dialog
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Excel Dosyası Seç",
                "",
                "Excel Dosyaları (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
            
            # Read Excel file with explicit engine and specify the sheet name
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name='Misafir Kayıtları')
            
            # Normalize column names by stripping whitespace and converting to a consistent format
            df.columns = [self.normalize_column_name(col) for col in df.columns]
            print(f"Excel columns after normalization: {df.columns.tolist()}") # Debugging
            
            # Validate required columns (normalized)
            expected_columns = list(self._REQUIRED_EXCEL_COLUMNS_MAP.keys())
            print(f"Expected required columns: {expected_columns}") # Debugging
            missing_normalized_columns = [col for col in expected_columns if col not in df.columns]
            if missing_normalized_columns:
                # Map back to original names for the error message
                missing_original_names = [self._REQUIRED_EXCEL_COLUMNS_MAP[col] for col in missing_normalized_columns]
                QMessageBox.warning(
                    self,
                    'Hata',
                    f'Excel dosyasında beklenen sütunlardan bazıları eksik veya yanlış isimlendirilmiş: {", ".join(missing_original_names)}.\n\nLütfen Excel şablonunu (Excel Şablonu İndir butonu ile edinebilirsiniz) kullanarak VERİLERİNİZİ DOĞRU SÜTUN ADLARIYLA girin. Sütun adları şablondakiyle BİREBİR AYNI olmalıdır.'
                )
                return
            
            # Track success/failure
            success_count = 0
            error_count = 0
            skipped_rows_details = []

            for index, row in df.iterrows():
                row_num = index + 2  # Excel row numbers start from 1, and header is row 1, so data starts from row 2
                try:
                    # Convert dates to string format and handle potential errors
                    check_in_date_str = str(row.get(self.normalize_column_name('Giriş Tarihi'), '')).strip()
                    check_out_date_str = str(row.get(self.normalize_column_name('Çıkış Tarihi'), '')).strip()

                    if not check_in_date_str or not check_out_date_str:
                        skipped_rows_details.append(f"Satır {row_num}: Giriş veya Çıkış Tarihi boş bırakılamaz.")
                        error_count += 1
                        continue

                    try:
                        check_in = pd.to_datetime(check_in_date_str, errors='coerce')
                        if pd.isna(check_in):
                            raise ValueError("Geçersiz tarih formatı")
                        check_in = check_in.strftime('%Y-%m-%d')
                    except Exception:
                        skipped_rows_details.append(f"Satır {row_num}: Geçersiz Giriş Tarihi formatı ('{check_in_date_str}').")
                        error_count += 1
                        continue
                    
                    try:
                        check_out = pd.to_datetime(check_out_date_str, errors='coerce')
                        if pd.isna(check_out):
                            raise ValueError("Geçersiz tarih formatı")
                        check_out = check_out.strftime('%Y-%m-%d')
                    except Exception:
                        skipped_rows_details.append(f"Satır {row_num}: Geçersiz Çıkış Tarihi formatı ('{check_out_date_str}').")
                        error_count += 1
                        continue

                    # Validate dates for logical consistency (check_out after check_in)
                    if not validate_dates(check_in, check_out):
                        skipped_rows_details.append(f"Satır {row_num}: Çıkış Tarihi, Giriş Tarihinden önce olamaz.")
                        error_count += 1
                        continue

                    # Convert hotel purchase price and handle potential errors
                    hotel_purchase_price_str = str(row.get(self.normalize_column_name('Otel Alış Fiyatı'), '')).strip()
                    if not hotel_purchase_price_str:
                        skipped_rows_details.append(f"Satır {row_num}: Otel Alış Fiyatı boş bırakılamaz.")
                        error_count += 1
                        continue
                    try:
                        hotel_purchase_price = float(hotel_purchase_price_str.replace(',', '.').replace('₺', '').replace(' ', ''))
                        if hotel_purchase_price <= 0:
                            raise ValueError("Sıfır veya negatif değer")
                    except Exception:
                        skipped_rows_details.append(f"Satır {row_num}: Geçersiz Otel Alış Fiyatı formatı ('{hotel_purchase_price_str}').")
                        error_count += 1
                        continue
                    
                    # Convert hotel sale price and handle potential errors
                    hotel_sale_price_str = str(row.get(self.normalize_column_name('Otel Satış Fiyatı'), '')).strip()
                    if not hotel_sale_price_str:
                        skipped_rows_details.append(f"Satır {row_num}: Otel Satış Fiyatı boş bırakılamaz.")
                        error_count += 1
                        continue
                    try:
                        hotel_sale_price = float(hotel_sale_price_str.replace(',', '.').replace('₺', '').replace(' ', ''))
                        if hotel_sale_price <= 0:
                            raise ValueError("Sıfır veya negatif değer")
                    except Exception:
                        skipped_rows_details.append(f"Satır {row_num}: Geçersiz Otel Satış Fiyatı formatı ('{hotel_sale_price_str}').")
                        error_count += 1
                        continue

                    # Ensure required text fields are not empty after stripping
                    guest_name = str(row.get(self.normalize_column_name('Adı Soyadı'), '')).strip()
                    guest_title = str(row.get(self.normalize_column_name('Unvan'), '')).strip()
                    company_name = str(row.get(self.normalize_column_name('Kurum / Cari'), '')).strip()
                    country = str(row.get(self.normalize_column_name('Ülke'), '')).strip()
                    city = str(row.get(self.normalize_column_name('Şehir'), '')).strip()
                    room_type = str(row.get(self.normalize_column_name('Oda Tipi'), '')).strip()
                    hotel_name = str(row.get(self.normalize_column_name('Otel Adı'), '')).strip()

                    required_fields = {
                        'Adı Soyadı': guest_name,
                        'Unvan': guest_title,
                        'Kurum / Cari': company_name,
                        'Ülke': country,
                        'Şehir': city,
                        'Oda Tipi': room_type,
                        'Otel Adı': hotel_name,
                        'Otel Satış Fiyatı': hotel_sale_price_str 
                    }

                    missing_fields = [k for k, v in required_fields.items() if not v]
                    if missing_fields:
                        skipped_rows_details.append(f"Satır {row_num}: Eksik veya boş alanlar: {', '.join(missing_fields)}.")
                        error_count += 1
                        continue

                    # Validate room type
                    valid_room_types = ['Single Oda', 'Double Oda', 'Triple Oda', 'Suit Oda', 'Aile Odası']
                    if room_type not in valid_room_types:
                        skipped_rows_details.append(f"Satır {row_num}: Geçersiz Oda Tipi ('{room_type}'). Geçerli tipler: {', '.join(valid_room_types)}")
                        error_count += 1
                        continue

                    # Create stay record
                    self.stay_model.create_stay(
                        guest_name, guest_title, company_name, country, city,
                        check_in, check_out, room_type,
                        hotel_purchase_price, hotel_name, hotel_sale_price
                    )
                    success_count += 1
                except Exception as e:
                    skipped_rows_details.append(f"Satır {row_num}: Beklenmeyen hata: {str(e)}")
                    error_count += 1

            if success_count > 0:
                # Reload table
                self.load_stays()
                
                # Show result message
                result_message = f'Başarıyla içe aktarılan: {success_count}\nToplam Hatalı Kayıt: {error_count}'
                if skipped_rows_details:
                    result_message += "\n\nAtlanan satırlar ve nedenleri:\n" + "\n".join(skipped_rows_details[:10]) # Show up to 10 details
                    if len(skipped_rows_details) > 10:
                        result_message += f"\n... ve {len(skipped_rows_details) - 10} adet daha."
                
                QMessageBox.information(
                    self,
                    'İçe Aktarma Tamamlandı',
                    result_message
                )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                'Hata',
                f'Excel dosyası içe aktarılırken genel bir hata oluştu: {str(e)}\n\nLütfen Excel şablonunu kullanarak doğru formatta veri girişi yapın.'
            )
    
    def download_excel_template(self):
        """Create and download Excel template file."""
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Misafir Kayıtları'
            
            # Write headers
            headers = [
                "Adı Soyadı", "Unvan", "Kurum / Cari", "Ülke", "Şehir",
                "Giriş Tarihi", "Çıkış Tarihi", "Oda Tipi", "Otel Alış Fiyatı",
                "Otel Adı", "Otel Satış Fiyatı"
            ]
            ws.append(headers)

            # Set column widths for better readability
            column_widths = {
                "Adı Soyadı": 20,
                "Unvan": 15,
                "Kurum / Cari": 25,
                "Ülke": 15,
                "Şehir": 15,
                "Giriş Tarihi": 18,
                "Çıkış Tarihi": 18,
                "Oda Tipi": 15,
                "Otel Alış Fiyatı": 18,
                "Otel Adı": 25,
                "Otel Satış Fiyatı": 18
            }

            for i, header in enumerate(headers):
                ws.column_dimensions[get_column_letter(i+1)].width = column_widths.get(header, 15)

            # Add example data
            example_data = [
                ['Ahmet Yılmaz', 'Bay', 'ABC Şirketi', 'Türkiye', 'İstanbul', '2024-03-20', '2024-03-25', 'Single Oda', 500.00, 'Grand Hotel', 600.00],
                ['Ayşe Demir', 'Bayan', 'XYZ Holding', 'Türkiye', 'Ankara', '2024-03-21', '2024-03-23', 'Double Oda', 750.00, 'City Resort', 850.00]
            ]
            for row_data in example_data:
                ws.append(row_data)

            # Create room types sheet first
            room_types = ['Single Oda', 'Double Oda', 'Triple Oda', 'Suit Oda', 'Aile Odası']
            room_types_sheet = wb.create_sheet(title='Oda Tipleri')
            
            # Add room types to the sheet
            for i, room_type in enumerate(room_types, 1):
                room_types_sheet.cell(row=i, column=1, value=room_type)
            
            # Create named range for room types
            room_types_range = f"'{room_types_sheet.title}'!$A$1:$A${len(room_types)}"
            wb.defined_names.add(DefinedName('room_types', attr_text=room_types_range))

            # Add data validation for Room Type
            room_type_dv = DataValidation(type="list", formula1="=room_types", allow_blank=True)
            ws.add_data_validation(room_type_dv)
            room_type_dv.add('H2:H1048576')  # Apply to Room Type column (H)

            # Add data validation for Dates
            date_dv = DataValidation(type="date", operator="greaterThanOrEqual", formula1=QDate(2023, 1, 1).toString("yyyy-MM-dd"),
                                     errorStyle='stop',
                                     errorTitle='Geçersiz Tarih',
                                     error='Lütfen geçerli bir tarih girin (YYYY-MM-DD).')
            ws.add_data_validation(date_dv)
            date_dv.add('F2:G1048576')  # Apply to Giriş Tarihi (F) and Çıkış Tarihi (G)
            
            # Add data validation for Hotel Purchase Price (numeric and positive)
            purchase_price_dv = DataValidation(type="whole", operator="greaterThan", formula1="0",
                                     errorStyle="stop",
                                     errorTitle="Geçersiz Fiyat",
                                     error="Lütfen 0'dan büyük bir sayı girin.")
            ws.add_data_validation(purchase_price_dv)
            purchase_price_dv.add('I2:I1048576')  # Apply to Otel Alış Fiyatı (I)

            # Add data validation for Hotel Sale Price (numeric and positive)
            sale_price_dv = DataValidation(type="whole", operator="greaterThan", formula1="0",
                                     errorStyle="stop",
                                     errorTitle="Geçersiz Fiyat",
                                     error="Lütfen 0'dan büyük bir sayı girin.")
            ws.add_data_validation(sale_price_dv)
            sale_price_dv.add('K2:K1048576')  # Apply to Otel Satış Fiyatı (K)

            # Add instructions
            instructions = [
                'Excel Şablonu Kullanım Talimatları:',
                '',
                '1. Tüm sütunları doldurunuz:',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["adi_soyadi"]}: Misafirin tam adı',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["unvan"]}: Bay/Bayan',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["kurum_cari"]}: Kurum / Cari',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["ulke"]}: Misafirin ülkesi',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["sehir"]}: Misafirin şehri',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["giris_tarihi"]}: YYYY-MM-DD formatında (örn: 2024-03-20)',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["cikis_tarihi"]}: YYYY-MM-DD formatında (örn: 2024-03-25)',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["oda_tipi"]}: Dropdown menüden seçiniz',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["otel_alis_fiyati"]}: Sayısal değer (örn: 500.00)',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["otel_adi"]}: Otel Adı',
                f'   - {self._REQUIRED_EXCEL_COLUMNS_MAP["otel_satis_fiyati"]}: Sayısal değer (örn: 600.00)',
                '',
                '2. Örnek kayıtları silip kendi kayıtlarınızı ekleyebilirsiniz.',
                '3. Tarihleri YYYY-MM-DD formatında girdiğinizden emin olun.',
                '4. Oda tipini sağdaki dropdown menüden seçiniz.',
                '5. Otel alış fiyatını geçerli bir sayı olarak giriniz.',
                '6. Otel satış fiyatını geçerli bir sayı olarak giriniz.',
                '',
                'Not: Bu şablonu doldurduktan sonra "Excel\'den İçe Aktar" butonu ile verileri sisteme aktarabilirsiniz.'
            ]
            
            # Create instructions sheet
            instructions_sheet = wb.create_sheet(title='Kullanım Talimatları')
            for i, instruction in enumerate(instructions, 1):
                instructions_sheet.cell(row=i, column=1, value=instruction)
            
            # Set column width for instructions
            instructions_sheet.column_dimensions['A'].width = 80

            # Hide the room types sheet
            room_types_sheet.sheet_state = 'hidden'
            
            # Set active sheet to Misafir Kayıtları
            wb.active = wb['Misafir Kayıtları']
            
            # Save the workbook
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Şablonunu Kaydet",
                "misafir_kayit_sablonu.xlsx",
                "Excel Dosyaları (*.xlsx)"
            )
            
            if not file_path:
                return

            # Add .xlsx extension if not present
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

            wb.save(file_path)
            QMessageBox.information(
                self,
                'Başarılı',
                f'Excel şablonu başarıyla oluşturuldu:\n{file_path}'
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                'Hata',
                f'Excel şablonu oluşturulurken hata oluştu: {str(e)}'
            )
    
    def closeEvent(self, event):
        """Handle window close event."""
        self.db.close()
        event.accept()

    def apply_filters(self):
        """Apply filters to the table by reloading stays with current filter values."""
        guest_name = self.filter_guest_name_input.text().strip()
        country = self.filter_country_input.text().strip()
        city = self.filter_city_input.text().strip()
        room_type = self.filter_room_type_combo.currentText()
        
        self.load_stays(guest_name, country, city, room_type, self.current_sort_column, self.current_sort_order)

    def sort_data(self, column_index: int):
        """Sort the table data based on the clicked column.
        
        Args:
            column_index (int): The index of the column that was clicked.
        """
        if column_index == self.current_sort_column:
            self.current_sort_order = Qt.DescendingOrder if self.current_sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            self.current_sort_column = column_index
            self.current_sort_order = Qt.AscendingOrder
        
        self.apply_filters() # Reapply filters and sorting

    def export_excel(self):
        """Export data from the table to an Excel file."""
        try:
            # Fetch all stays to export
            stays = self.stay_model.get_all_stays(sort_column="ID", sort_order="ASC")
            
            if not stays:
                QMessageBox.information(self, "Bilgi", "Aktarılacak kayıt bulunmamaktadır.")
                return

            # Create a Pandas DataFrame from the stays data
            # Calculate 'Konaklama Süresi (Gece)' before creating DataFrame
            processed_stays = []
            for stay in stays:
                check_in = datetime.strptime(stay['check_in_date'], "%Y-%m-%d")
                check_out = datetime.strptime(stay['check_out_date'], "%Y-%m-%d")
                nights = (check_out - check_in).days
                stay_with_nights = stay.copy()
                stay_with_nights['nights_stayed'] = nights # Use a temporary key for internal use
                processed_stays.append(stay_with_nights)

            df = pd.DataFrame(processed_stays)
            
            # Select and rename columns for the Excel export
            df = df[[
                'id', 'guest_name', 'guest_title', 'company_name', 'country', 'city',
                'check_in_date', 'check_out_date', 'room_type',
                'hotel_purchase_price', 'hotel_purchase_total_amount', 'hotel_name', 'hotel_sale_price', 'total_sale_amount', 'nights_stayed'
            ]]
            df.columns = [
                'ID', 'Adı Soyadı', 'Unvan', 'Kurum / Cari', 'Ülke', 'Şehir',
                'Giriş Tarihi', 'Çıkış Tarihi', 'Oda Tipi',
                'Otel Alış Fiyatı', 'Otel Alış Toplam Ücreti', 'Otel Adı', 'Otel Satış Fiyatı', 'Toplam Satış Tutarı', 'Konaklama Süresi (Gece)'
            ]
            
            # Prompt user to save the file
            file_path, _ = QFileDialog.getSaveFileName(self, "Excel\'e Aktar", "konaklama_kayitlari.xlsx", "Excel Files (*.xlsx)")
            
            if file_path:
                df.to_excel(file_path, index=False) # Export to Excel
                QMessageBox.information(self, "Başarılı", f"Veriler başarıyla Excel\'e aktarıldı:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel\'e aktarılırken bir hata oluştu: {e}")

    def generate_puantaj_report(self):
        """Generate detailed stay report in Excel format."""
        try:
            # Fetch all stays
            stays = self.stay_model.get_detailed_stay_report()
            
            if not stays:
                self.show_info_message("Puantaj raporu oluşturulacak kayıt bulunmamaktadır.")
                return

            # Prepare data for the report
            report_data = []
            for stay in stays:
                print(f"Processing stay with keys: {stay.keys()}") # Added for debugging
                report_data.append({
                    'ID': stay['id'],
                    'Adı Soyadı': stay['guest_name'],
                    'Unvan': stay['guest_title'],
                    'Kurum / Cari': stay['company_name'],
                    'Ülke': stay['country'],
                    'Şehir': stay['city'],
                    'Giriş Tarihi': stay['check_in_date'],
                    'Çıkış Tarihi': stay['check_out_date'],
                    'Oda Tipi': stay['room_type'],
                    'Otel Alış Fiyatı': stay['hotel_purchase_price'],
                    'Otel Alış Toplam Ücreti': stay['hotel_purchase_total_amount'],
                    'Otel Adı': stay['hotel_name'],
                    'Otel Satış Fiyatı': stay['hotel_sale_price'],
                    'Toplam Satış Tutarı': stay['total_sale_amount'],
                    'Konaklama Süresi (Gece)': stay['nights'],
                    'Toplam Misafir Sayısı': stay['total_guests'],
                    'Toplam Konaklama Sayısı': stay['total_stays'],
                    'Toplam Konaklama Günü': stay['total_nights'],
                    'Toplam Alış Geliri': stay['total_revenue_purchase'],
                    'Toplam Satış Geliri': stay['total_revenue_sale']
                })
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(report_data)
            
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Puantaj Raporunu Kaydet",
                "konaklama_puantaj_raporu.xlsx",
                "Excel Dosyaları (*.xlsx)"
            )
            
            if not file_path:
                return

            # Add .xlsx extension if not present
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

            df.to_excel(file_path, index=False)
            self.report_generated.emit() # Emit signal after report generation
            QMessageBox.information(
                self,
                'Başarılı',
                f'Puantaj raporu başarıyla oluşturuldu:\n{file_path}'
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                'Hata',
                f'Puantaj raporu oluşturulurken hata oluştu: {str(e)}'
            )

    def show_info_message(self, message: str = "Rapor başarıyla oluşturuldu!"):
        """Display an informational message box."""
        QMessageBox.information(self, 'Bilgi', message)